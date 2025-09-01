import os
import time
import random
import pandas as pd
import concurrent.futures
import logging
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import *
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from functools import lru_cache  # 新增：用于缓存


# 颜色代码定义
class Color:
    RESET = "\033[0m"
    RED = "\033[91m"
    GREEN = "\033[92m"
    YELLOW = "\033[93m"
    BLUE = "\033[94m"
    PURPLE = "\033[95m"
    CYAN = "\033[96m"
    BOLD = "\033[1m"


def print_status(message, color=Color.RESET, end='\n'):
    """带颜色的状态输出"""
    print(f"{color}{message}{Color.RESET}", end=end)


@lru_cache(maxsize=None)  # 新增：缓存用户代理列表
def get_user_agents():
    """获取用户代理列表（缓存避免重复创建）"""
    return [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.212 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.114 Safari/537.36"
    ]


def create_driver():
    """创建无头Chrome浏览器驱动，优化启动参数"""
    # 禁用Selenium的日志输出
    logging.getLogger('selenium').setLevel(logging.CRITICAL)

    chrome_options = Options()
    # 核心优化参数
    chrome_options.add_argument("--headless=new")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument("--log-level=3")
    chrome_options.add_argument("--silent")

    # 新增：性能优化参数
    chrome_options.add_argument("--disable-extensions")  # 禁用扩展
    chrome_options.add_argument("--disable-plugins")  # 禁用插件
    chrome_options.add_argument("--disable-popup-blocking")  # 禁用弹窗拦截
    chrome_options.add_argument("--disable-images")  # 禁用图片加载（大幅提升速度）
    chrome_options.add_argument("--blink-settings=imagesEnabled=false")  # 进一步禁用图片
    chrome_options.add_argument("--disable-background-networking")  # 禁用后台网络请求
    chrome_options.add_argument("--no-first-run")  # 跳过首次运行设置
    chrome_options.add_argument("--no-default-browser-check")  # 跳过默认浏览器检查

    # 随机User-Agent
    chrome_options.add_argument(f"user-agent={random.choice(get_user_agents())}")

    # 禁用性能日志
    chrome_options.set_capability("goog:loggingPrefs", {"performance": "OFF", "browser": "OFF"})

    # 初始化驱动
    try:
        service = Service(
            executable_path="./chromedriver.exe",
            service_log_path=os.devnull
        )
        driver = webdriver.Chrome(service=service, options=chrome_options)
        driver.set_page_load_timeout(15)  # 超时时间从30秒减少到15秒
        driver.set_script_timeout(10)  # 新增：脚本超时设置
        return driver
    except Exception as e:
        print_status(f"创建浏览器驱动失败: {str(e)}", Color.RED)
        return None


def extract_info(driver, url):
    """优化信息提取逻辑，减少不必要操作"""
    result = {
        "url": url,
        "Code": "",
        "Message": "",
        "Resource": "",
        "RequestId": "",
        "valid": True,
        "access_denied": False
    }

    try:
        # 减少随机延迟范围（从1-3秒改为0.5-1.5秒）
        time.sleep(random.uniform(0.5, 1.5))

        driver.get(url)
        page_source = driver.page_source.lower()

        # 先检查关键错误状态，快速返回
        if "not exist" in page_source:
            result["valid"] = False
            print_status(f"❌ 无效 URL: {url}", Color.RED)
            return result

        if "access denied" in page_source:
            result["valid"] = False
            result["access_denied"] = True
            print_status(f"🚫 访问拒绝: {url}", Color.YELLOW)
            return result

        print_status(f"✅ 有效 URL: {url}", Color.GREEN)

        # 优化XPath提取逻辑，使用更高效的选择器
        xpath_map = {
            "Code": "//*[contains(text(), 'Code')]/following-sibling::*",
            "Message": "//*[contains(text(), 'Message')]/following-sibling::*",
            "Resource": "//*[contains(text(), 'Resource')]/following-sibling::*",
            "RequestId": "//*[contains(text(), 'RequestId')]/following-sibling::*"
        }

        # 批量提取信息，减少重复代码
        for key, xpath in xpath_map.items():
            try:
                element = driver.find_element("xpath", xpath)
                result[key] = element.text.strip()
            except:
                result[key] = "N/A"

    except TimeoutException:
        result["Message"] = "Timeout"
        result["valid"] = False
        print_status(f"⏱️ 超时 URL: {url}", Color.YELLOW)
    except WebDriverException as e:
        result["Message"] = f"Error: {str(e)}"
        result["valid"] = False
        print_status(f"⚠️ 错误 URL: {url} ({str(e)})", Color.YELLOW)
    except Exception as e:
        result["Message"] = f"Unexpected error: {str(e)}"
        result["valid"] = False
        print_status(f"⚠️ 异常 URL: {url} ({str(e)})", Color.YELLOW)

    return result


# 新增：驱动池管理类，复用浏览器实例
class DriverPool:
    def __init__(self, max_drivers=5):
        self.max_drivers = max_drivers
        self.drivers = []
        self._initialize_drivers()

    def _initialize_drivers(self):
        for _ in range(self.max_drivers):
            driver = create_driver()
            if driver:
                self.drivers.append(driver)

    def get_driver(self):
        if not self.drivers:
            return create_driver()
        return self.drivers.pop()

    def return_driver(self, driver):
        if len(self.drivers) < self.max_drivers:
            self.drivers.append(driver)
        else:
            driver.quit()

    def close_all(self):
        for driver in self.drivers:
            driver.quit()
        self.drivers = []


def process_url(url, driver_pool):
    """使用驱动池处理单个URL，复用浏览器实例"""
    driver = driver_pool.get_driver()
    if not driver:
        return None

    try:
        return extract_info(driver, url)
    finally:
        if driver:
            driver_pool.return_driver(driver)


def create_unique_filename(base_name, extension):
    """创建不重复的文件名"""
    counter = 1
    filename = f"{base_name}.{extension}"

    while os.path.exists(filename):
        filename = f"{base_name}_{counter}.{extension}"
        counter += 1

    return filename


def format_excel(file_path):
    """优化Excel格式化逻辑"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path)
        ws = wb.active

        # 设置样式（保持不变）
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 格式化标题行
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # 优化数据行处理：只处理有值的单元格
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                if cell.value is not None:  # 只处理有值的单元格
                    cell.border = thin_border
                    cell.alignment = Alignment(wrap_text=True)

        # 调整列宽
        column_widths = {
            "A": 8,  # 序号
            "B": 40,  # URL
            "C": 10,  # Code
            "D": 30,  # Message
            "E": 20,  # Resource
            "F": 30  # RequestId
        }

        for col in column_widths:
            ws.column_dimensions[col].width = column_widths[col]

        wb.save(file_path)
        return True
    except Exception as e:
        print_status(f"格式化Excel时出错: {str(e)}", Color.RED)
        return False


def main():
    print_status("\n" + "=" * 60, Color.CYAN)
    print_status(f"{Color.BOLD}                      URL批量检测工具                      {Color.RESET}", Color.CYAN)
    print_status(f"{Color.BOLD}                     (OSS URL Checker)                    {Color.RESET}", Color.CYAN)
    print_status("=" * 60 + "\n", Color.CYAN)

    # 读取URL列表
    try:
        with open("url.txt", "r", encoding="utf-8") as f:
            urls = [line.strip() for line in f if line.strip()]
    except FileNotFoundError:
        print_status("错误: 未找到url.txt文件", Color.RED)
        return
    except Exception as e:
        print_status(f"读取url.txt时出错: {str(e)}", Color.RED)
        return

    total_urls = len(urls)
    if total_urls == 0:
        print_status("url.txt文件中没有URL", Color.YELLOW)
        return

    print_status(f"发现 {total_urls} 个URL，开始检测...", Color.BLUE)
    print_status("-" * 60, Color.CYAN)

    # 优化并发策略：根据URL数量动态调整线程数
    max_workers = min(10, total_urls) if total_urls > 0 else 1
    driver_pool = DriverPool(max_drivers=max_workers)  # 创建驱动池

    results = []
    processed = 0

    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        # 使用驱动池处理URL
        futures = {executor.submit(process_url, url, driver_pool): url for url in urls}

        for future in concurrent.futures.as_completed(futures):
            processed += 1
            percentage = (processed / total_urls) * 100
            bar_length = 50
            filled_length = int(bar_length * processed // total_urls)
            bar = '█' * filled_length + '-' * (bar_length - filled_length)
            print_status(f'\r检测进度: |{bar}| {percentage:.1f}% ({processed}/{total_urls})', Color.BLUE, end='')

            try:
                results.append(future.result())
            except Exception as e:
                print_status(f"\n处理URL时出错: {str(e)}", Color.RED)

        print()

    # 关闭所有驱动
    driver_pool.close_all()

    # 结果统计（保持不变）
    valid_results = [r for r in results if r and r["valid"] and not r["access_denied"]]
    invalid_results = [r for r in results if r and not r["valid"] and not r["access_denied"]]
    access_denied_count = len([r for r in results if r and r["access_denied"]])
    error_results = len([r for r in results if r is None])

    valid_count = len(valid_results)
    invalid_count = len(invalid_results) + error_results

    print_status("\n" + "-" * 60, Color.CYAN)
    print_status(f"{Color.BOLD}检测结果统计:{Color.RESET}", Color.PURPLE)
    print_status(f"总检测URL数: {total_urls}", Color.BLUE)
    print_status(f"有效URL数: {valid_count} {Color.GREEN}✅{Color.RESET}", Color.GREEN)
    print_status(f"无效URL数: {invalid_count} {Color.RED}❌{Color.RESET}", Color.RED)
    print_status(f"访问拒绝URL数: {access_denied_count} {Color.YELLOW}🚫{Color.RESET}", Color.YELLOW)
    print_status("-" * 60, Color.CYAN)

    if valid_count == 0:
        print_status("没有有效的URL可写入Excel", Color.YELLOW)
        return

    # 准备写入Excel的数据
    for i, result in enumerate(valid_results, 1):
        result["序号"] = i

    df = pd.DataFrame(valid_results)
    df = df[["序号", "url", "Code", "Message", "Resource", "RequestId"]]

    excel_file = create_unique_filename("result", "xlsx")

    try:
        df.to_excel(excel_file, index=False)
        print_status(f"\n有效URL信息已保存到: {excel_file}", Color.GREEN)

        if format_excel(excel_file):
            print_status("Excel文件已美化完成", Color.GREEN)
    except Exception as e:
        print_status(f"保存Excel时出错: {str(e)}", Color.RED)


if __name__ == "__main__":
    main()