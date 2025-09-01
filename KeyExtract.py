import requests
from bs4 import BeautifulSoup
import re
import time
import os
import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


# 颜色代码定义
class Colors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'


def print_separator():
    """打印分隔线美化输出"""
    print(f"\n{Colors.OKBLUE}" + "=" * 60 + f"{Colors.ENDC}\n")


def create_filename_from_url(url):
    """从URL生成安全的文件名，包含域名和路径信息"""
    # 提取域名部分
    domain_match = re.search(r'https?://([^/]+)', url)
    if not domain_match:
        return "unknown_domain_unknown_path"

    domain = domain_match.group(1)
    # 替换域名中的特殊字符为下划线
    safe_domain = re.sub(r'[^\w]', '_', domain)

    # 提取路径中的最后一部分
    path_parts = url.split('/')[3:]  # 跳过协议和域名部分
    # 过滤空字符串并获取最后一个有效路径段
    valid_path_parts = [part for part in path_parts if part.strip()]
    last_path = valid_path_parts[-1] if valid_path_parts else "unknown_path"

    # 替换路径中的特殊字符为下划线
    safe_path = re.sub(r'[^\w]', '_', last_path)

    # 组合域名和路径部分
    return f"{safe_domain}_{safe_path}"


def get_unique_filename(base_name):
    """生成唯一Excel文件名，若存在则添加递增序号"""
    full_base = f"{base_name}_result"

    # 检查文件是否存在
    if not os.path.exists(f"{full_base}.xlsx"):
        return f"{full_base}.xlsx"

    # 若存在则添加序号
    counter = 1
    while True:
        # 格式化序号为两位数字（01, 02, ..., 99）
        numbered_name = f"{full_base}_{counter:02d}.xlsx"
        if not os.path.exists(numbered_name):
            return numbered_name
        counter += 1
        # 限制最大序号，避免无限循环
        if counter > 99:
            return f"{full_base}_{counter}.xlsx"


def format_excel(file_path):
    """美化Excel表格"""
    # 打开Excel文件
    df = pd.read_excel(file_path)
    writer = pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace')
    df.to_excel(writer, index=False, sheet_name='Results')

    # 获取工作表
    worksheet = writer.sheets['Results']

    # 定义样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    alignment = Alignment(vertical='center', wrap_text=False)

    # 设置表头样式
    for cell in worksheet[1]:  # 表头在第二行（索引1）
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = alignment

    # 设置数据单元格样式和列宽
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = alignment

    # 特殊设置序号列居中显示
    if '序号' in [cell.value for cell in worksheet[1]]:
        index_col = [cell.value for cell in worksheet[1]].index('序号')
        index_letter = worksheet.cell(row=1, column=index_col + 1).column_letter
        for cell in worksheet[index_letter]:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 自动调整列宽
    for column_cells in worksheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells) + 2
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(length, 50)  # 最大宽度限制

    writer.close()


def extract_and_process():
    # 美化欢迎界面
    print(f"\n{Colors.HEADER}" + "*" * 60)
    print(" " * 15 + "URL标签提取与Excel生成工具 v1.0")
    print("*" * 60 + f"{Colors.ENDC}")

    # 获取用户输入的URL
    url = input(f"\n{Colors.BOLD}请输入要访问的URL: {Colors.ENDC}").strip()

    # 显示处理中状态
    print(f"\n{Colors.OKBLUE}正在处理，请稍候...{Colors.ENDC}", end="", flush=True)

    try:
        # 发送请求访问URL
        response = requests.get(url, timeout=10)
        response.raise_for_status()  # 检查请求是否成功

        # 解析XML内容
        soup = BeautifulSoup(response.text, 'xml')

        # 提取所有Key标签的内容 - 作为主要数据列表
        key_tags = soup.find_all('Key')
        if not key_tags:
            print("\r" + " " * 30 + "\r", end="")  # 清除"处理中"提示
            print_separator()
            print(f"{Colors.WARNING}⚠️  未找到任何<Key>标签内容{Colors.ENDC}")
            print_separator()
            return

        # 要提取的标签列表
        tags_to_extract = ['Key', 'Size', 'Type', 'ID', 'LastModified']

        # 收集所有数据
        data = []
        total_items = len(key_tags)

        # 基础URL用于拼接完整链接
        base_url_match = re.search(r'(https?://[^/]+/)', url)
        base_url = base_url_match.group(1) if base_url_match else ""

        for i, key_tag in enumerate(key_tags):
            # 进度提示
            progress = (i + 1) / total_items * 100
            print(f"\r{Colors.OKBLUE}正在处理: {progress:.1f}%{Colors.ENDC}", end="", flush=True)

            item = {}
            # 添加序号列（从1开始）
            item['序号'] = i + 1

            # 提取Key内容
            item['Key'] = key_tag.get_text(strip=True)

            # 生成完整链接作为Host
            item['Host'] = f"{base_url}{item['Key']}" if base_url else item['Key']

            # 提取其他标签（如果存在）
            for tag in tags_to_extract[1:]:  # 跳过已经处理的Key
                # 尝试查找当前Key标签后的同级标签
                tag_element = key_tag.find_next_sibling(tag)
                # 如果找不到，尝试在整个文档中查找
                if not tag_element:
                    tag_element = soup.find(tag)
                if tag_element:
                    item[tag] = tag_element.get_text(strip=True)

            data.append(item)

        # 生成基础文件名（包含域名和路径信息）
        base_filename = create_filename_from_url(url)

        # 获取唯一Excel文件名
        excel_filename = get_unique_filename(base_filename)

        # 创建DataFrame并保存为Excel
        df = pd.DataFrame(data)
        # 确保序号列在最前面
        if '序号' in df.columns:
            cols = ['序号'] + [col for col in df.columns if col != '序号']
            df = df[cols]

        df.to_excel(excel_filename, index=False, sheet_name='Results')

        # 美化Excel表格
        format_excel(excel_filename)

        # 美化输出结果
        print("\r" + " " * 30 + "\r", end="")  # 清除进度提示
        print_separator()
        print(f"{Colors.OKGREEN}✅  处理完成！{Colors.ENDC}")
        print(f"{Colors.BOLD}" + "-" * 40 + f"{Colors.ENDC}")
        print(f"📊  结果已保存至：{Colors.UNDERLINE}{excel_filename}{Colors.ENDC}")
        print(f"{Colors.BOLD}" + "-" * 40 + f"{Colors.ENDC}")

        # 显示提取的列信息
        columns_info = f"提取的列: {', '.join(df.columns.tolist())}"
        print(f"{Colors.OKBLUE}{columns_info}{Colors.ENDC}")
        print(f"{Colors.OKBLUE}📊  统计信息：共提取 {total_items} 条记录{Colors.ENDC}")
        print_separator()

    except requests.exceptions.RequestException as e:
        print("\r" + " " * 30 + "\r", end="")  # 清除进度提示
        print_separator()
        print(f"{Colors.FAIL}❌  访问URL时出错: {str(e)[:50]}...{Colors.ENDC}")
        print_separator()
    except Exception as e:
        print("\r" + " " * 30 + "\r", end="")  # 清除进度提示
        print_separator()
        print(f"{Colors.FAIL}❌  处理过程中出错: {str(e)[:50]}...{Colors.ENDC}")
        print_separator()


if __name__ == "__main__":
    extract_and_process()
